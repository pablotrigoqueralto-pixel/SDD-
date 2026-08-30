"""Tabular file reader: CSV (sniffed delimiter/encoding) and .xlsx via openpyxl.

Headers map case- and accent-insensitively against canonical names and their
Spanish aliases, so the user's real Sage/Excel files load without editing.
"""

import csv
import io
import unicodedata
from collections.abc import Mapping

from openpyxl import load_workbook

from app.domain.shared.errors import FieldError, ValidationFailedError

MAX_ROWS = 2000
MAX_BYTES = 5 * 1024 * 1024
_XLSX_MAGIC = b"PK\x03\x04"


def _file_error(message: str, code: str) -> ValidationFailedError:
    errors: list[FieldError] = [{"field": "file", "message": message, "code": code}]
    return ValidationFailedError(errors)


def _normalise_header(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value.strip().lower())
    return "".join(char for char in decomposed if not unicodedata.combining(char))


def _decode(content: bytes) -> str:
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return content.decode("cp1252")


def _csv_rows(content: bytes) -> list[list[str]]:
    text = _decode(content)
    first_line = text.splitlines()[0] if text.splitlines() else ""
    delimiter = ";" if first_line.count(";") >= first_line.count(",") else ","
    return [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _xlsx_rows(content: bytes) -> list[list[str]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise _file_error(
            "The file could not be read as an Excel workbook", "file_unreadable"
        ) from exc
    sheet = workbook.worksheets[0]
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell) for cell in row])
        if len(rows) > MAX_ROWS + 1:
            break
    workbook.close()
    return rows


def read_table(
    filename: str,
    content: bytes,
    *,
    columns: Mapping[str, tuple[str, ...]],
    required: frozenset[str],
) -> list[dict[str, str]]:
    """Rows as {canonical: trimmed value}; raises ValidationFailedError (422) on file problems."""
    if len(content) > MAX_BYTES:
        raise _file_error("The file exceeds 5 MB; split it and import in parts", "file_too_large")
    is_xlsx = filename.lower().endswith(".xlsx") or content.startswith(_XLSX_MAGIC)
    try:
        raw_rows = _xlsx_rows(content) if is_xlsx else _csv_rows(content)
    except ValidationFailedError:
        raise
    except Exception as exc:
        raise _file_error("The file could not be read", "file_unreadable") from exc
    if not raw_rows:
        raise _file_error("The file has no header row", "file_empty")

    alias_to_canonical = {
        _normalise_header(alias): canonical
        for canonical, aliases in columns.items()
        for alias in aliases
    }
    header_map: dict[int, str] = {}
    for index, header in enumerate(raw_rows[0]):
        canonical = alias_to_canonical.get(_normalise_header(header))
        if canonical is not None and canonical not in header_map.values():
            header_map[index] = canonical
    missing = sorted(required - set(header_map.values()))
    if missing:
        raise _file_error(f"Missing required columns: {', '.join(missing)}", "missing_columns")

    data_rows = raw_rows[1:]
    rows: list[dict[str, str]] = []
    for raw in data_rows:
        values = {canonical: "" for canonical in columns}
        for index, canonical in header_map.items():
            if index < len(raw):
                values[canonical] = raw[index].strip()
        if any(values.values()):
            rows.append(values)
        if len(rows) > MAX_ROWS:
            raise _file_error(
                f"The file has more than {MAX_ROWS} data rows; split it and import in parts",
                "too_many_rows",
            )
    return rows
